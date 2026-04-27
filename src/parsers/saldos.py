import re
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from src.cleaning.normalize import normalize_spaces, to_number, to_str_id


def _cell(row: Tuple[Any, ...], idx: int) -> Any:
    return row[idx] if idx < len(row) else None


def _parse_file_year(file_path: str) -> int:
    if "26" in file_path:
        return 2026
    return 2025


def _extract_cutoff_date(ws) -> Optional[str]:
    date_pattern = re.compile(r"\b(\d{2}/\d{2}/\d{4})\b")
    found: List[datetime] = []

    for row in ws.iter_rows(min_row=1, max_row=25, max_col=20, values_only=True):
        for value in row:
            text = normalize_spaces(value)
            if not text:
                continue
            for match in date_pattern.findall(text):
                try:
                    found.append(datetime.strptime(match, "%d/%m/%Y"))
                except ValueError:
                    pass

    if not found:
        return None
    return max(found).date().isoformat()


def parse_saldos(file_path: str) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    records: List[Dict[str, Any]] = []
    rejected: List[Dict[str, Any]] = []

    current_cuenta = ""
    current_cuenta_nombre = ""
    current_record: Optional[Dict[str, Any]] = None
    year = _parse_file_year(file_path)
    cutoff = _extract_cutoff_date(ws)
    if not cutoff:
        cutoff = "2026-04-30" if year == 2026 else "2025-12-31"

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        a = normalize_spaces(_cell(row, 0))
        b = normalize_spaces(_cell(row, 1))
        c = normalize_spaces(_cell(row, 2))

        if a.lower().startswith("cuenta") and b:
            match = re.match(r"(\d+)[\s\-]*(.*)", b)
            if match:
                current_cuenta = to_str_id(match.group(1))
                current_cuenta_nombre = normalize_spaces(match.group(2))
            else:
                current_cuenta = ""
                current_cuenta_nombre = b
            current_record = None
            continue

        tercero_id = to_str_id(_cell(row, 0))
        if tercero_id.isdigit() and c and current_cuenta:
            saldo_anterior = to_number(_cell(row, 6)) or 0.0
            debitos = to_number(_cell(row, 9)) or 0.0
            creditos = to_number(_cell(row, 12)) or 0.0
            saldo_actual = to_number(_cell(row, 14))

            if saldo_actual is None:
                saldo_actual = saldo_anterior + debitos - creditos

            current_record = {
                "cuenta": current_cuenta,
                "nombre_cuenta": current_cuenta_nombre,
                "tercero_id": tercero_id,
                "tercero_nombre": c,
                "saldo_anterior": saldo_anterior,
                "debitos": debitos,
                "creditos": creditos,
                "saldo_actual": saldo_actual,
                "fecha_corte": cutoff,
                "anio": year,
                "fila_origen": row_idx,
            }
            records.append(current_record)
            continue

        if current_record and not a and c:
            has_numeric = any(to_number(_cell(row, idx)) is not None for idx in [6, 9, 12, 14])
            if not has_numeric:
                current_record["tercero_nombre"] = normalize_spaces(
                    f"{current_record.get('tercero_nombre', '')} {c}"
                )
                continue

        if tercero_id and not tercero_id.isdigit() and c and current_cuenta:
            rejected.append(
                {
                    "fuente": file_path,
                    "hoja": ws.title,
                    "fila": row_idx,
                    "motivo": "tercero_id_no_numerico",
                }
            )

    return records, rejected
