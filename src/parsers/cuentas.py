from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook

from src.cleaning.normalize import normalize_spaces, to_str_id


def parse_cuentas(file_path: str) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    records: List[Dict[str, Any]] = []
    rejected: List[Dict[str, Any]] = []

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cuenta = to_str_id(row[0] if len(row) > 0 else None)
        nombre = normalize_spaces(row[1] if len(row) > 1 else None)
        nivel = row[2] if len(row) > 2 else None

        if not cuenta or not cuenta[0].isdigit():
            continue
        if not nombre:
            rejected.append(
                {
                    "fuente": file_path,
                    "hoja": ws.title,
                    "fila": row_idx,
                    "motivo": "cuenta_sin_nombre",
                }
            )
            continue

        try:
            nivel_int = int(nivel) if nivel is not None else None
        except Exception:
            nivel_int = None

        records.append(
            {
                "cuenta": cuenta,
                "nombre_cuenta": nombre,
                "nivel": nivel_int,
            }
        )

    return records, rejected
