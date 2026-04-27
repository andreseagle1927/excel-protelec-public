from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from src.cleaning.normalize import (
    excel_serial_to_datetime,
    normalize_spaces,
    to_number,
    to_str_id,
)


def _cell(row: Tuple[Any, ...], idx: int) -> Any:
    return row[idx] if idx < len(row) else None


def _is_operation_row(row: Tuple[Any, ...]) -> bool:
    no_oper = _cell(row, 0)
    fecha = _cell(row, 1)
    tipo_doc = normalize_spaces(_cell(row, 3))
    no_doc = normalize_spaces(_cell(row, 6))
    if not tipo_doc or "Cuenta" in tipo_doc:
        return False
    if tipo_doc.isdigit():
        return False
    if not no_doc:
        return False
    return to_str_id(no_oper).isdigit() and (excel_serial_to_datetime(fecha) is not None)


def _is_detail_header(row: Tuple[Any, ...]) -> bool:
    a = normalize_spaces(_cell(row, 0)).lower()
    b = normalize_spaces(_cell(row, 1)).lower()
    return "cuenta" in a and "tercero" in b


def _is_detail_item_row(row: Tuple[Any, ...]) -> bool:
    cuenta = to_str_id(_cell(row, 0))
    tercero = to_str_id(_cell(row, 1))
    return cuenta.isdigit() and tercero.isdigit()


def _parse_file_year(file_path: str) -> int:
    if "2026" in file_path:
        return 2026
    return 2025


def parse_movimientos(file_path: str) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    records: List[Dict[str, Any]] = []
    rejected: List[Dict[str, Any]] = []
    year = _parse_file_year(file_path)

    current_operation: Dict[str, Any] = {}
    current_item: Optional[Dict[str, Any]] = None

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if _is_operation_row(row):
            op_date = excel_serial_to_datetime(_cell(row, 1))
            current_operation = {
                "no_oper": to_str_id(_cell(row, 0)),
                "fecha": op_date.date().isoformat() if isinstance(op_date, datetime) else None,
                "tipo_doc": normalize_spaces(_cell(row, 3)),
                "no_doc": to_str_id(_cell(row, 6)),
                "clasif": normalize_spaces(_cell(row, 8)),
                "detalle_operacion": normalize_spaces(_cell(row, 9)),
                "anio": year,
            }
            current_item = None
            continue

        if _is_detail_header(row):
            current_item = None
            continue

        if _is_detail_item_row(row) and current_operation:
            cuenta = to_str_id(_cell(row, 0))
            tercero_id = to_str_id(_cell(row, 1))
            centro_costos = to_str_id(_cell(row, 3))
            detalle = normalize_spaces(_cell(row, 5))
            valor_base = to_number(_cell(row, 9))
            debito = to_number(_cell(row, 12))
            credito = to_number(_cell(row, 14))

            if debito is None and credito is None:
                rejected.append(
                    {
                        "fuente": file_path,
                        "hoja": ws.title,
                        "fila": row_idx,
                        "motivo": "item_sin_valor",
                    }
                )
                current_item = None
                continue

            current_item = {
                **current_operation,
                "cuenta": cuenta,
                "tercero_id": tercero_id,
                "centro_costos": centro_costos,
                "detalle": detalle,
                "valor_base": valor_base or 0.0,
                "debito": debito or 0.0,
                "credito": credito or 0.0,
                "fila_origen": row_idx,
            }
            records.append(current_item)
            continue

        if current_item:
            extra_detalle = normalize_spaces(_cell(row, 5))
            if extra_detalle:
                current_item["detalle"] = normalize_spaces(
                    f"{current_item.get('detalle', '')} {extra_detalle}"
                )

    return records, rejected
